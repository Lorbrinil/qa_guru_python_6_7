from openpyxl import load_workbook
import os.path

from conftest import resources_path


# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_xlsx():
    workbook = load_workbook(os.path.join(resources_path, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    sheet_value = sheet.cell(row=3, column=2).value
    print(sheet_value)

    assert sheet_value == 'Mara'
